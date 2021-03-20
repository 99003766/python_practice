import openpyxl
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

excel_file = Workbook()
#import openpyxl
#from openpyxl import Workbook

excel_file = Workbook()
wb = openpyxl.load_workbook('studentinfo.xlsx')
sheets = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']
excel_sheet = excel_file.create_sheet(title='MasterSheet11', index=0)
n=int(input("number of persons:" ))


for g in range(1, n+1):

    print("enter", g, " person information")
    xin = int(input("enter ps number: "))
    yin = input("enter name: ")
    zin = input("enter mailid: ")

    t = 1

    for sheet in sheets:
        sh = wb[sheet]  # Get a sheet from the workbook.
        max_r = sh.max_row
        max_c = sh.max_column
        if t <= 10:
            for r in range(1, max_r + 1):
                if sh.cell(row=r, column=1).value == xin and sh.cell(row=r, column=2).value == yin and sh.cell(row=r, column=3).value == zin:
                    print("Check in ExcelFile: ")
                    for c in range(1, max_c + 1):
                        if g==1:
                            str1 = 'A' + str(t)

                            str2 = 'B' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
                        else:
                            str1 = chr(67+g) + str(t)

                            str2 = chr(68+g) + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value


        else:
            for r in range(4, max_r + 1):
                if sh.cell(row=r, column=1).value == xin and sh.cell(row=r, column=2).value == yin and sh.cell(row=r,
                                                                                                               column=3).value == zin:
                    for c in range(4, max_c + 1):
                        if g==1:

                            str1 = 'A' + str(t)

                            str2 = 'B' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
                        else:
                            str1 = chr(67+g) + str(t)

                            str2 = chr(68+g) + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value

    excel_file.save(filename="final.xlsx")
#for ploting bar graph

wb = openpyxl.load_workbook('Final.xlsx')

excel_file = wb
# Get workbook active sheet
# from the active attribute.
sheet = wb['MasterSheet11']

# write o to 9 in 1st column of the active sheet
# create data for plotting
values = Reference(sheet, min_col=2, min_row=9,
                   max_col=sheet.max_column, max_row=17)

# Create object of BarChart class
chart = BarChart()

# adding data to the Bar chart object
chart.add_data(values)

# set the title of the chart
chart.title = " BAR-CHART "

# set the title of the x-axis
chart.x_axis.title = " X_AXIS "

# set the title of the y-axis
chart.y_axis.title = " Y_AXIS "

# add chart to the sheet
# the top-left corner of a chart
# is anchored to cell E2 .
sheet.add_chart(chart, "E2")

# save the file
wb.save("Final.xlsx")
