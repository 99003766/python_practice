import pandas as pd
# import numpy as np
from openpyxl import load_workbook
# import openpyxl
excel_file = pd.ExcelFile('Excel.xlsx')

df_first = pd.read_excel(excel_file, sheet_name=0)
df_second = pd.read_excel(excel_file, sheet_name=1)
df_third = pd.read_excel(excel_file, sheet_name=2)
df_fourth = pd.read_excel(excel_file, sheet_name=3)
df_fifth = pd.read_excel(excel_file, sheet_name=4)

f3 = df_first.merge(df_second,on="Official Email Address",how="left")
f4 = f3.merge(df_third,on="Official Email Address", how="left")
f5 = f4.merge(df_fourth,on="Official Email Address", how="left")
#f6 = f5.merge(df_fifth,on="Official Email Address", how="left")
Input = input("Official Email Address")
df_Input = pd.DataFrame(f5, columns=['PS NUMBER', 'Display Name', 'Official Email Address', 'Training Hall', 'Floor Number', 'Date Of Joining',	'Domain', 'Attending Genesis', 'System Number', 'Team Number',
                                    'BUS NUMBER', 'Working Hours', 'Marks Subject1', 'Marks Subject2', 'Marks Subject3',
                                    'Stream', 'Address', 'Area', 'Room Number', 'Permanent Address', 'Data1', 'Data2'
                                    ])
#df_Input = pd.DataFrame(f5, columns=['Official Email Address'])
df_Input.set_index('Official Email Address', inplace=True)
result = df_Input.loc[Input]
print(result)
path = "Excel.xlsx"
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine='openpyxl')
writer.book = book
if 'MasterSheet' in book.sheetnames:
    ref = book['MasterSheet']
    book.remove(ref)
result.to_excel(writer, sheet_name='MasterSheet')
writer.save()
writer.close()
