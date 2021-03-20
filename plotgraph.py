import openpyxl
import pandas as pd
from openpyxl import Workbook
from matplotlib import pyplot as plt
excel_file = Workbook()
#import openpyxl
#from openpyxl import Workbook

excel_file = Workbook()
wb = openpyxl.load_workbook('studentinfo.xlsx')
sheets = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']
excel_sheet = excel_file.create_sheet(title='MasterSheet11', index=0)
n=int(input("number of persons:" ))


for g in range(1, n+1):

    print("enter", g, " person information")
    xin = int(input("enter ps number: "))
    yin = input("enter name: ")
    zin = input("enter mailid: ")

    t = 1

    for sheet in sheets:
        sh = wb[sheet]  # Get a sheet from the workbook.
        max_r = sh.max_row
        max_c = sh.max_column
        if t <= 10:
            for r in range(1, max_r + 1):
                if sh.cell(row=r, column=1).value == xin and sh.cell(row=r, column=2).value == yin and sh.cell(row=r, column=3).value == zin:
                    print("Check in ExcelFile: ")
                    for c in range(1, max_c + 1):
                        if g==1:
                            str1 = 'A' + str(t)

                            str2 = 'B' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
                        else:
                            str1 = 'E' + str(t)

                            str2 = 'F' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value


        else:
            for r in range(4, max_r + 1):
                if sh.cell(row=r, column=1).value == xin and sh.cell(row=r, column=2).value == yin and sh.cell(row=r,
                                                                                                               column=3).value == zin:
                    for c in range(4, max_c + 1):
                        if g==1:

                            str1 = 'A' + str(t)

                            str2 = 'B' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value
                        else:
                            str1 = 'E' + str(t)

                            str2 = 'F' + str(t)
                            t = t + 1
                            excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                            excel_sheet[str2] = sh.cell(row=r, column=c).value

    excel_file.save(filename="final.xlsx")

#ploting the bar graph
sh = wb['Sheet2']
l=[]
p=[]
t=1
max_r=sh.max_row
max_c=sh.max_column
for r in range(2,max_r+1):
    l.append(sh.cell(row=r, column=7).value)
for r in range(2,max_r+1):
    p.append(sh.cell(row=r, column=8).value)
global_num=pd.DataFrame(p,l)
print(global_num)
str1='A'+str(t)
t=t+1
dates = pd.to_datetime(global_num.index)
plt.plot(global_num)
plt.show()
plt.savefig('plot.xlsx')

"""
enter name: Jayasimha Reddy Ganapuram
enter mailid: jayasimha.ganapuram@ltts.com
Check in ExcelFile: 
enter 2  person information
enter ps number: 99003737
enter name: Jeshwanth Kumar Ega
enter mailid: jeshwanth.ega@ltts.com
"""
