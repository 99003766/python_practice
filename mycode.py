import pandas as pd
from openpyxl import load_workbook
#j=str(input("enter name"))
#k = str(input("enter email id"))
n=int(input("enter ps number"))
p=input("name: ")
q=input("gmail: ")
sheet_name=['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']
s=pd.read_excel('studentinfo.xlsx', sheet_name)
y=s['Sheet1']
y=y[y['PS_NUMBER'] == n]
y=y[y['Official Email Address'] == q]
y=y[y['Display Name'] == p]
df = pd.DataFrame(y, columns=['PS_NUMBER','Display Name','Official Email Address'])

for i in s.keys():
    x=s[i]
    t = x[x['PS_NUMBER']==n]
    t = x[x['Official Email Address'] == q]
    t = x[x['Display Name'] == p]
    col = x.columns
    for j in col:
        df[j]=t[j]
#df.to_excel('pythonexcel1.xlsx',sheet_name='master',index=False)
path = "studentinfo.xlsx"
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine='openpyxl')
writer.book = book
# searching for the master sheet in all sheets
if 'MasterSheet' in book.sheetnames:
    ref = book['MasterSheet']
    # removing the previous data in the sheet
    book.remove(ref)
    # writing the data into the master sheet
df.to_excel(writer, sheet_name='MasterSheet')
# saving  the data in the master sheet
writer.save()
writer.close()
#enter nameRishab Pankajkumar Ostawal
#enter email idrishab.ostawal@ltts.com
#enter ps number99003754