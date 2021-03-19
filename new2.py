import openpyxl
from openpyxl import Workbook
excel_file = Workbook()
wb = openpyxl.load_workbook('Excel.xlsx')
sheets=['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']
excel_sheet = excel_file.create_sheet(title='MasterSheet1', index=0)
xin = input()
t=1
for sheet in sheets:
    sh = wb[sheet]  # Get a sheet from the workbook.
    max_r = sh.max_row
    max_c = sh.max_column
        for r in range(1, max_r + 1):
            if sh.cell(row=r, column=3).value == xin:
                for c in range(1, max_c + 1):
                    str1 = 'A' + str(t)

                    str2 = 'B' + str(t)
                    t = t + 1
                    excel_sheet[str1] = str(sh.cell(row=1, column=c).value)
                    excel_sheet[str2] = sh.cell(row=r, column=c).value



excel_file.save(filename="Excel.xlsx")
