import pandas as pd
from openpyxl import load_workbook

df_total = pd.DataFrame()

sheets = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']

yin = input("Enter Official Email Address")

for sheet in sheets:

    y = pd.read_excel('studentinfo.xlsx', sheet_name=sheet)
    df1 = pd.DataFrame(y[y['Official Email Address'] == yin], columns=y.columns)

    df_total = df_total.join(df1, how='outer', lsuffix='left', rsuffix='right')


path = "studentinfo.xlsx"
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine='openpyxl')
writer.book = book

if 'MasterSheet' in book.sheetnames:
    ref = book['MasterSheet']

    book.remove(ref)

df_total.to_excel(writer, sheet_name='MasterSheet')

writer.save()
writer.close()