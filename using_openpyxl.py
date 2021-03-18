from openpyxl import load_workbook
path = "Excel.xlsx"
wb_obj = load_workbook(path)
sheet_obj = wb_obj.active
#cell_obj = sheet_obj.cell(row = 5, column = 5)
#print(cell_obj.value)
#print(sheet_obj.max_row)
#print(sheet_obj.max_column)
#max_col = sheet_obj.max_column

# Loop will print all columns name
#for i in range(1, max_col):
#    cell_obj = sheet_obj.cell(row=1, column=i)
#    print(cell_obj.value)
#m_row = sheet_obj.max_row

# Loop will print all values
# of first column
#for i in range(1, m_row + 1):
#    cell_obj = sheet_obj.cell(row=i, column=1)
#    print(cell_obj.value)
#max_col = sheet_obj.max_column

# Will print a particular row value
#for i in range(1, max_col + 1):
#    cell_obj = sheet_obj.cell(row=2, column=i)
#    print(cell_obj.value, end=" ")





